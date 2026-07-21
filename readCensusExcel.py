# readCensusExcel.py - Tabulates county population and census tracts

import openpyxl
import pprint

print("Opening workbook...")

wb = openpyxl.load_workbook("censuspopdata.xlsx")

sheet = wb["Population by Census Tract"]

county_data = {}

print(f"Worksheet: {sheet.title}")
print(f"Rows: {sheet.max_row}")
print(f"Columns: {sheet.max_column}")

print("Reading rows...")

for row in range(2, sheet.max_row + 1):
    state = sheet["B" + str(row)].value
    county = sheet["C" + str(row)].value
    pop = sheet["D" + str(row)].value

    county_data.setdefault(state, {})

    print(county_data)
    break